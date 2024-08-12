from flask import (Flask, render_template, request, redirect, url_for ) # type: ignore
import os
import json
import openpyxl # type: ignore
from werkzeug.utils import secure_filename # type: ignore

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads/'

# Ensure directories exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs('data', exist_ok=True)

# Register (Account Creation)
@app.route('/register', methods=['POST'])
def register():
    username = request.form['username']
    password = request.form['password']
    
    users_path = './data/users.json'
    users = []
    
    if os.path.exists(users_path):
        with open(users_path, 'r') as file:
            users = json.load(file)
    
    users.append({'username': username, 'password': password})
    
    with open(users_path, 'w') as file:
        json.dump(users, file, indent=2)
    
    return redirect(url_for('login'))

# Login
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        users_path = './data/users.json'
        
        if os.path.exists(users_path):
            with open(users_path, 'r') as file:
                users = json.load(file)
            
            for user in users:
                if user['username'] == username and user['password'] == password:
                    return redirect(url_for('admission_form'))
        
        return "Invalid credentials!"
    
    return render_template('login.html')

# Admission Form
@app.route('/admission_form')
def admission_form():
    return render_template('admission.html')

@app.route('/submit_form', methods=['POST'])
def submit_form():
    form_data = {
        'name': request.form['name'],
        'contact': request.form['contact'],
        'father_name': request.form['father_name'],
        'mother_name': request.form['mother_name'],
        'address': request.form['address'],
        'fees_paid': request.form['fees_paid'],
        'total_amount': request.form['total_amount'],
        'balance_amount': request.form['balance_amount'],
        'due_date': request.form['due_date'],
        'parent_contact': request.form['parent_contact']
    }
    
    # Handling file uploads
    id_proof = request.files['id_proof']
    marksheet = request.files['marksheet']
    
    form_data['id_proof'] = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(id_proof.filename))
    form_data['marksheet'] = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(marksheet.filename))
    
    id_proof.save(form_data['id_proof'])
    marksheet.save(form_data['marksheet'])
    
    # Save form data to Excel
    file_path = './data/admissions.xlsx'
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
    else:
        workbook = openpyxl.Workbook()
    
    sheet = workbook.active
    sheet.title = 'Admissions'
    
    # If the sheet is empty, add headers
    if sheet.max_row == 1:
        headers = ['Name', 'Contact', 'Father\'s Name', 'Mother\'s Name', 'Address', 'ID Proof', 'Marksheet', 'Fees Paid', 'Total Amount', 'Balance Amount', 'Due Date', 'Parent Contact']
        sheet.append(headers)
    
    data = [
        form_data['name'],
        form_data['contact'],
        form_data['father_name'],
        form_data['mother_name'],
        form_data['address'],
        form_data['id_proof'],
        form_data['marksheet'],
        form_data['fees_paid'],
        form_data['total_amount'],
        form_data['balance_amount'],
        form_data['due_date'],
        form_data['parent_contact']
    ]
    
    sheet.append(data)
    workbook.save(file_path)
    
    return "Form submitted successfully!"

if __name__ == '__main__':
    app.run(debug=True)